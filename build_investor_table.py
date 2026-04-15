import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ========== SHEET 1: INVERSIONISTAS PRIORITARIOS ==========
ws = wb.active
ws.title = "VCs Prioritarios GoSafe"

header_fill = PatternFill('solid', fgColor='1a1a2e')
header_font = Font(name='Arial', bold=True, color='c8ff00', size=11)
high_fill = PatternFill('solid', fgColor='e8f5e9')
med_fill = PatternFill('solid', fgColor='fff3e0')
low_fill = PatternFill('solid', fgColor='e3f2fd')
corp_fill = PatternFill('solid', fgColor='fce4ec')
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc')
)

headers = [
    "Prioridad", "Fondo / Entidad", "Persona de Contacto (Startco)",
    "Cargo / Rol", "Stage", "Verticales Clave",
    "Fit con GoSafe", "Website", "Acci\u00f3n Recomendada"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

data = [
    # HIGH PRIORITY - Healthtech + IA directo
    ["ALTA", "Krealo", "Gianfranco Smarrelli / Pedro M. Rubio Solf / Ivan Bautista", "Investment Team", "Seed / Serie A", "Healthtech, Insurtech, B2B, Infra Digital", "Healthtech + B2B + LATAM. Brazo de Credicorp.", "https://krealo.pe/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "EWA Capital", "Mar\u00eda Jos\u00e9 Vinueza", "Investment Team", "Pre-seed a Series A", "Educaci\u00f3n, Salud, Servicios Financieros", "Salud como vertical directa. Stage ideal.", "https://ewa.capital/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "BID Lab", "Carolina Jim\u00e9nez / Daniel P\u00e9rez", "Investment Officers", "Fondos + Directo", "Salud/Biotech, Edtech, Climatetech", "Salud/Biotech expl\u00edcito. Capital + credibilidad institucional.", "https://bidlab.org/es", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "IGNIA Partners", "Fabrice Serfati", "Partner", "Temprana a Crecimiento", "Healthtech, SaaS, Fintech, Tech aplicada", "Healthtech expl\u00edcito. Multi-stage, puede acompa\u00f1ar varias rondas.", "https://www.ignia.vc/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Capria Ventures", "Manuela V\u00e1squez", "Investment Team", "Seed / Series A+", "Healthtech, IA aplicada, SaaS, Agtech", "Doble fit: Healthtech + IA aplicada. El m\u00e1s alineado.", "https://capria.vc/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "M Capital Syndicate", "Andr\u00e9s Fl\u00f3rez Llano", "GP / Lead", "Preseed / Seed", "AI, Health/BioTech, Fintech, Climate", "IA + Health + Colombia. Stage preseed/seed perfecto.", "LinkedIn: M Capital", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Latin Leap", "Stefan Krautwald / Nancy Zapata", "Partners", "Seed / Early-stage", "Healthtech, Fintech, Edtech, Proptech", "Healthtech + seed + LATAM. Ideal para escalamiento.", "https://latinleap.vc/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Alive Ventures", "Laura Mu\u00f1oz / Mar\u00eda Camila Vernaza", "Investment Team", "Colombia & Per\u00fa", "Healthtech, Edtech, Agtech, Circular Economy", "Healthtech + Colombia + impacto social.", "http://www.alive-ventures.com/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Cube Ventures", "David Ortiz Mu\u00f1oz / Zindy Amador", "Investment Team", "Pre-seed / Seed", "Healthtech, Fintech, Edtech, Smart Cities", "Healthtech + infra tech. Stage alineado.", "https://www.cube.ventures/en", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "New Ventures Capital", "Matteo Gilliotti / Manuela Bedoya", "Investment Team", "Early-stage / Seed", "Salud, Fintech, Educaci\u00f3n, Clean-tech, Impacto", "Salud + impacto social. Narrativa poderosa para GoSafe.", "https://www.nvgroup.org/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Simma Capital", "Sara Pati\u00f1o / Daniel Bland\u00f3n", "Investment Team", "Pre-seed / Seed", "Healthtech, Software, Fintech, Edtech", "Healthtech + software SaaS. Encaja perfecto.", "LinkedIn: Simma Capital", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Lotux VC", "Mat Gantar", "Partner", "Preseed ($50K-$100K)", "Health, Finance, Insurance, Education", "Health + ticket peque\u00f1o. Co-inversi\u00f3n ideal.", "http://www.lotux.vc/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Corteza Capital", "Felipe Uribe", "Partner", "Pre-seed a Serie A", "AI (GPU Native), Fintech, Cybersecurity", "IA como core thesis. \u00c1ngulo tech-first.", "https://www.cortezacapital.com/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ALTA", "Caribe Ventures", "Anuar Harb", "GP", "Microfondo (2do sem 2026)", "AI pre-seed ($500K fondo)", "100% AI. Sembrar relaci\u00f3n ahora.", "https://www.caribe.ventures/", "Networking + seguimiento post-evento"],
    ["ALTA", "CAF", "Diana Buitrago", "Investment Officer", "Institucional", "Desarrollo, Innovaci\u00f3n, Impacto LATAM", "Banca de desarrollo. Capital + credibilidad + escala LATAM.", "https://www.caf.com/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],

    # MEDIUM - Agn\u00f3sticos / SaaS / IA general
    ["MEDIA", "Winter Kpital", "Sergio Zuluaga / Mauricio Jim\u00e9nez", "Partners", "Pre-seed a Growth", "SaaS, AI, Marketplaces, Enterprise", "SaaS + AI. GoSafe como enterprise SaaS m\u00e9dico.", "https://winterkpital.com/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Inqlab", "Andr\u00e9s Fonnegra", "Partner", "Early-stage / Seed", "Agn\u00f3stico (healthtech recurrente)", "Healthtech en radar. Aceleradora colombiana.", "https://www.inqlab.co/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "30N Ventures", "Salvador", "Investment Team", "Pre-Seed / Seed", "Fintech, SaaS, Marketplaces, Consumer Tech", "SaaS B2B angle.", "https://www.30n.vc/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "B Venture Capital", "Valentina Bojac\u00e1", "Investment Team", "Pre-seed / Seed", "Fintech, SaaS, Agn\u00f3stico Tech", "Agn\u00f3stico + SaaS. Para completar ronda.", "https://www.bventure.capital/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Zetta Ventures", "Alejandra L\u00f3pez De Mesa / Amalia Escobar", "Investment Team", "Pre-seed / Seed", "SaaS, Marketplaces, B2B, B2C", "SaaS B2B. Fondo colombiano bien conectado.", "https://www.zettaventures.co/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Axon Partners Group", "Sof\u00eda Areiza / Diego Serna", "LATAM Team", "Growth / Multi-stage", "Technology, Deep Tech, AI, Sustainability", "AI + Tech. M\u00e1s para rondas futuras (growth).", "https://axonpartnersgroup.com/", "Networking + relaci\u00f3n largo plazo"],
    ["MEDIA", "Angel Hub / AngelHub Ventures", "Jose Luis Cimental / Mario Garc\u00eda D\u00e1vila / Marcelo Garcia", "Partners", "Seed / Temprana", "Agn\u00f3stico", "Red de \u00e1ngeles M\u00e9xico. Co-inversi\u00f3n + expansi\u00f3n MX.", "https://www.angelhub.mx/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Digital Hub Monterrey", "Mauricio Mu\u00f1oz", "Partner", "Seed / Pre-Series A", "Startups de IA (multicorporativo)", "100% IA. Puerta a M\u00e9xico corporativo.", "http://www.mtydigitalhub.com/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Viwala", "Cristian Latorre / Cindy Diakit\u00e9 / Andrea Galeano", "Investment Team", "Pre-seed / Seed", "Fintech, Impacto Social, PYMES", "Impacto social + PYMES. \u00c1ngulo de acceso a salud.", "https://viwala.com/", "Pitch en showcase + seguimiento"],
    ["MEDIA", "Lanchmon", "Boris Daniel Lancheros / Valentina Gaviria", "Founders", "Plataforma/Ecosistema", "Conecta startups con inversores y corporativos", "Plataforma de visibilidad. Puede amplificar alcance.", "https://lanchmon.com/", "Registrarse en plataforma + networking"],

    # ESTRAT\u00c9GICO - Corporate VCs
    ["ESTRAT\u00c9GICO", "Bancolombia Ventures", "Carlos Arturo Figueroa Fadul", "Venture Capital", "CVC", "Corporate VC del banco m\u00e1s grande de Colombia", "Distribuci\u00f3n + capital. Red hospitalaria.", "https://www.bancolombia.com/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ESTRAT\u00c9GICO", "EPM Ventures", "Jaime Rios / In\u00e9s Fanery Gonz\u00e1lez", "Venture Team", "CVC ($150MM COP)", "Ciudades inteligentes, Innovaci\u00f3n", "Infra salud Medell\u00edn. Tickets grandes.", "https://www.epm.com.co/", "Agendar reuni\u00f3n 1-on-1 en VC Week"],
    ["ESTRAT\u00c9GICO", "Organizaci\u00f3n Corona CVC", "Melissa Escobar Restrepo", "CVC Team", "CVC", "Fit con core business Sodimac Corona", "Bienestar + hogar. Conexi\u00f3n indirecta.", "https://empresa.corona.co/", "Networking en evento"],
    ["ESTRAT\u00c9GICO", "Progreso X", "Rodrigo Zetina", "CVC Team", "CVC", "Open Innovation arm de Progreso", "Innovaci\u00f3n corporativa. Bienestar empresarial.", "https://www.progreso-x.com/", "Networking en evento"],
    ["ESTRAT\u00c9GICO", "Satrack CVC", "Estefan\u00eda Arboleda / Delia Var\u00f3n", "CVC Team", "Pre-seed / Seed", "Transporte, Movilidad, Log\u00edstica", "Menor fit directo. IoT/data angle.", "https://satrackventures.co/", "Networking general"],
    ["ESTRAT\u00c9GICO", "Renault-Sofasa", "Ana Mar\u00eda Posada / Daniel Osorio Ochoa", "Innovation Team", "Corporate", "Movilidad, Innovaci\u00f3n", "Corporate innovaci\u00f3n. Bienestar laboral angle.", "https://www.renault.co/", "Networking general"],
    ["ESTRAT\u00c9GICO", "Ruta N", "Carlos Eduardo Gonz\u00e1lez Almeida", "Investment/Innovation", "Institucional", "Innovaci\u00f3n Medell\u00edn", "Facilitador de pilotos en Medell\u00edn. Cr\u00edtico.", "https://rutanmedellin.org/", "Agendar reuni\u00f3n + pedir intro a hospitales"],

    # ADICIONALES con potencial
    ["OPORTUNIDAD", "Inga Capital", "Oscar Gualdron", "GP", "VC", "Varios", "Nuevo fondo. Explorar tesis.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Linku Ventures", "Daniel Toro Uribe / Felipe Cano / Sebasti\u00e1n Aristiz\u00e1bal", "Partners", "VC", "Varios", "Fondo colombiano. Explorar fit.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "El Hub Ventures", "Julian Arango", "Partner", "VC", "Varios", "Venture fund. Explorar tesis.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Pacific Ventures", "Leonardo Borrero", "Partner", "VC", "Varios", "Fondo Pac\u00edfico colombiano.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Startup Grind / Ganas Ventures", "Jos\u00e9 Alejandro \u00c1lvarez Calero", "Partner", "VC", "Varios", "Red global Startup Grind + fondo.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Kamay Ventures", "Laura Prado", "Investment Team", "VC", "Varios", "Explorar tesis de inversi\u00f3n.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "inVerita", "Melissa M\u00fanera / Eduardo Monsalvo", "Partners", "VC", "Varios", "Fondo de inversi\u00f3n. Explorar fit.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Alborada Ventures", "Jorge De La Hoz", "Partner", "VC", "Varios", "Nuevo fondo. Explorar tesis.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Bioracer", "Jaime Gallego / Juan Lotero", "Partners", "VC", "BioTech / Health", "BioTech. Posible fit con salud.", "N/A", "Pitch r\u00e1pido + seguimiento"],
    ["OPORTUNIDAD", "Pecap", "Paloma Aparicio L\u00e1inez Lozada", "Director", "Asociaci\u00f3n", "Capital Semilla Per\u00fa", "Red de fondos peruanos. Puerta a Per\u00fa.", "https://www.pecap.pe/", "Networking en evento"],
    ["OPORTUNIDAD", "Confrapar", "Pablo Angel", "Partner", "VC Brasil", "Tecnolog\u00eda", "Fondo brasile\u00f1o tech. Puerta a Brasil.", "https://confrapar.com.br/", "Networking en evento"],
    ["OPORTUNIDAD", "Petra Secondaries", "Fabricio Zabala", "Partner", "Serie A+", "PE Generalista", "M\u00e1s para rondas grandes futuras.", "https://petrasecondaries.com/", "Relaci\u00f3n a largo plazo"],
    ["OPORTUNIDAD", "Alaya Capital", "Luis Narro", "Partner", "Pre-seed a Series A", "Climate-tech, Impact, Energy, Sustainability", "\u00c1ngulo impacto + sostenibilidad en salud.", "http://www.alaya-capital.com/", "Pitch en showcase"],
    ["OPORTUNIDAD", "Mundi Ventures", "Mar\u00eda Isabel Mu\u00f1oz", "Investment Team", "Pre-seed a Series B", "Insurtech, Smart Cities, Proptech, Fintech", "Insurtech. Conexi\u00f3n con aseguradoras de salud.", "https://mundiventures.com/es/", "Pitch en showcase"],
    ["OPORTUNIDAD", "Eatable Adventures", "Miguel \u00c1ngel Fl\u00f3rez", "Partner", "VC FoodTech", "FoodTech / AgTech", "Bajo fit directo. Networking general.", "N/A", "Networking general"],
    ["OPORTUNIDAD", "JEC Capital Group", "Sof\u00eda Upegui", "Investment Team", "VC", "Varios", "Explorar tesis.", "LinkedIn: JEC Capital", "Networking en evento"],
    ["OPORTUNIDAD", "Vibras Lab", "Daniel Vieco Calder\u00f3n", "Partner", "Family Office / VC", "PE, Entertainment, Innovation", "Family Office Karol G / J Balvin. Alto perfil.", "https://vibraslab.com", "Networking VIP"],
    ["OPORTUNIDAD", "Tronex", "Jos\u00e9 Jaime Parra / Fede V\u00e1squez", "Corporate", "Corporate", "Tecnolog\u00eda / Energ\u00eda", "Corporate. Exploratorio.", "N/A", "Networking general"],
    ["OPORTUNIDAD", "Grupo Bios", "Felipe Gutierrez / Milton Duque", "Corporate Innovation", "Corporate", "Agroindustria / Innovaci\u00f3n", "Corporate agro. Bienestar laboral angle.", "N/A", "Networking general"],
    ["OPORTUNIDAD", "Industrias Haceb", "Andr\u00e9s Felipe Restrepo Cuartas", "Innovation", "Corporate", "Innovaci\u00f3n / Manufactura", "Corporate. Exploratorio.", "N/A", "Networking general"],
    ["OPORTUNIDAD", "Orbit Ventures", "Daniel Osorio", "Partner", "VC", "Varios", "Explorar tesis.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Impulsa Venture Capital", "Morelia Quintero", "Partner", "VC Venezuela/LATAM", "Varios", "Fondo venezolano. Explorar.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Total Holding Family Office", "Camilo Mej\u00eda", "Family Office", "Family Office", "Varios", "Family Office. Inversi\u00f3n directa posible.", "N/A", "Networking VIP"],
    ["OPORTUNIDAD", "Boostart", "Giancarlo Braccia Avila", "Partner", "Venture Studio", "Venture Studio", "Venture Studio. Apoyo operativo posible.", "https://www.boostart.vc/", "Networking en evento"],
    ["OPORTUNIDAD", "Vertical Partners / VerticalLabs", "William Otero / Salom\u00e9 \u00c1lvarez / Renata Velez", "Partners", "VC", "Varios", "Fondo + Laboratorio. Explorar.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "adn.vc (ADN.VC)", "Nicol\u00e1s Di Pace", "Partner", "Pre-Seed", "FinTech, PropTech (MX, CL, CO, PE)", "Pre-seed LATAM. Si hay \u00e1ngulo fintech en salud.", "https://adn.vc/", "Pitch en showcase"],
    ["OPORTUNIDAD", "TaskUs", "Ruben Orozco / Gersson Nore\u00f1a", "Corporate", "Corporate", "Outsourcing / Tech", "Corporate tech. IA angle.", "N/A", "Networking general"],
    ["OPORTUNIDAD", "Tecnol\u00f3gico de Monterrey", "C\u00e9sar S\u00e1nchez", "Academic/VC", "Academic", "IA / Innovaci\u00f3n", "Conexi\u00f3n acad\u00e9mica + M\u00e9xico.", "N/A", "Networking"],
    ["OPORTUNIDAD", "Flink", "David Osorio / Ana Maria Cartagena", "Investment Team", "VC", "Varios", "Fintech/VC. Explorar.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "Alternativa Inversiones", "Tatiana Sof\u00eda Rinc\u00f3n", "Investment Team", "VC", "Varios", "Fondo alternativo. Explorar.", "N/A", "Networking en evento"],
    ["OPORTUNIDAD", "PeakU", "Santiago Gonz\u00e1lez", "VC/Platform", "VC", "Tech / Talento", "Tech talent platform. Exploratorio.", "N/A", "Networking en evento"],
]

fill_map = {"ALTA": high_fill, "MEDIA": med_fill, "ESTRAT\u00c9GICO": corp_fill, "OPORTUNIDAD": low_fill}

for i, row in enumerate(data, 2):
    fill = fill_map.get(row[0], low_fill)
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = fill
    ws.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True)

widths = [14, 28, 40, 18, 22, 38, 42, 32, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f"A1:I{len(data)+1}"
ws.freeze_panes = "A2"

# ========== SHEET 2: INVERSIONISTAS INDEPENDIENTES Y OTROS ==========
ws2 = wb.create_sheet("Inversionistas Independientes")

headers2 = ["Nombre Completo", "Empresa / Afiliaci\u00f3n", "Tipo", "Notas"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

independientes = [
    ["Jorge Andr\u00e9s Garc\u00eda Pava", "Independiente", "Angel Investor", "Inversionista independiente registrado en Startco"],
    ["John Restrepo", "Independiente", "Angel Investor", "Inversionista independiente"],
    ["Grace Torres", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Ana Prieto", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Humberto Ariza Hern\u00e1ndez", "N/A", "Angel Investor", "Inversionista independiente"],
    ["Luis Gonzalo Tob\u00f3n Gonz\u00e1lez", "Independiente", "Angel Investor", "Inversionista independiente"],
    ["Mauricio Conedera", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Andrea Manosalva Davies", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Adriana Herrera", "N/A", "Angel Investor", "Inversionista independiente"],
    ["Cecilia Salazar Jaramillo", "N/A", "Angel Investor", "Inversionista independiente"],
    ["Judith Posada Mar\u00edn", "N/A", "Angel Investor", "Inversionista independiente"],
    ["Manuel Arango", "N/A", "Angel Investor", "Inversionista independiente"],
    ["Alejandro Giraldo Estrada", "Independiente", "Angel Investor", "Inversionista independiente"],
    ["Mateo G\u00f3mez Zuluaga", "MCJ G\u00f3mez SAS", "Empresario/Inversor", "Empresa familiar con apetito de inversi\u00f3n"],
    ["Santiago Largo Cardona", "Largo Easy Corp", "Empresario/Inversor", "Corporaci\u00f3n con posible inter\u00e9s"],
    ["Jos\u00e9 Ignacio S\u00e1enz Hoyos", "Sego Company", "Empresario/Inversor", "Empresa con posible inter\u00e9s en salud"],
    ["Carlos Iglesias", "MCH Solutions Inc", "Empresario/Inversor", "Empresa tech. Posible inter\u00e9s en IA."],
    ["Maori David Herrera", "Dmaori SAS", "Empresario/Inversor", "Empresa local"],
    ["Santiago Cala", "Alfred", "Empresario/Inversor", "Plataforma tech"],
    ["Juan Saldarriaga", "360Lateral", "Empresario/Inversor", "Empresa tech/consulting"],
    ["Alejandro Gonz\u00e1lez", "Macca", "Empresario/Inversor", "Explorar inter\u00e9s"],
    ["Luis Salamanca", "Qubilo", "Empresario/Inversor", "Explorar inter\u00e9s"],
    ["Edwar Neira Mar\u00edn", "Escappy Travel", "Empresario/Inversor", "Travel tech"],
    ["Steven Cabrera", "Grupo Epark", "Empresario/Inversor", "Grupo empresarial"],
    ["Camilo Johnson", "Everfit Inc", "Empresario/Inversor", "Health/Fitness tech. Posible sinergia."],
    ["Jacobo Ortiz V\u00e9lez", "Orvel Construcciones", "Empresario/Inversor", "Construcci\u00f3n"],
    ["Shyra Morales Tapia", "M&P Angels / MAC Consultores", "Angel Network", "Red de \u00e1ngeles. Potencial co-inversi\u00f3n."],
    ["Juan Sim\u00f3n Ram\u00edrez Luna", "Supplies Industriales SAS", "Empresario/Inversor", "Industrial"],
    ["Esteban Roa", "Rise Autom\u00f3vil", "Empresario/Inversor", "Movilidad"],
    ["Jorge Correa", "Cohousing", "Empresario/Inversor", "Real estate tech"],
    ["Alexandra Baquero", "Sigma SAS", "Empresario/Inversor", "Empresa tech"],
    ["Mar\u00eda Alejandra Cano Arias", "Talento Consultores SAS", "Empresario/Inversor", "Consultor\u00eda talento"],
    ["Andr\u00e9s Felipe Sierra V\u00e1squez", "SM Digital SAS", "Empresario/Inversor", "Marketing digital"],
    ["Ruben Zapata", "Sampi Consultores", "Empresario/Inversor", "Consultor\u00eda"],
    ["Vanessa Carolina Rodr\u00edguez P\u00e9rez", "Inttel Go", "Empresario/Inversor", "Telecomunicaciones"],
    ["Carlos Betancur", "SPE SAS", "Empresario/Inversor", "Consultor\u00eda"],
    ["Manuel Felipe Tamayo D\u00edez", "Grupo ECB", "Empresario/Inversor", "Grupo empresarial"],
    ["Mateo Ruiz Espinosa", "Infovital", "Empresario/Inversor", "Posible sinergia salud"],
    ["Fabi\u00e1n Garc\u00eda", "iQuor", "Empresario/Inversor", "Tech"],
    ["Massimo Di Cesare Vielma", "N-iX", "Corporate", "Empresa tech global"],
    ["Ryan Ren", "Hytera", "Corporate", "Telecomunicaciones global"],
    ["Kevin Benegas", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Elmer Zapata", "Reputaci\u00f3n Online", "Empresario/Inversor", "Marketing digital"],
    ["Guillermo Alejandro Herrera Nassar", "VAAS Colombia SAS", "Empresario/Inversor", "Empresa local"],
    ["Rahul Ghosh", "Staaake", "Inversionista/Startup", "Posible inversor o founder"],
    ["Marlio Silva Barrero", "Industrias Molinillito / FoodTech", "Empresario/Inversor", "Agroindustria + FoodTech"],
    ["Ricardo Pinz\u00f3n D\u00edaz", "TechGrow Inversiones SAS", "VC/Angel", "Fondo peque\u00f1o de inversi\u00f3n tech"],
    ["Erick Ospina Puertas", "Coinsenda", "Startup/Inversor", "Crypto/Fintech. Posible \u00e1ngel."],
    ["Lady S\u00e1nchez Trujillo", "Andromeda", "Inversionista", "Explorar tesis"],
    ["Alejandro V\u00e9lez", "Debita", "Inversionista/Fintech", "Fintech. Explorar inter\u00e9s en health."],
    ["Camilo Tibocha", "Hesvor SAS", "Empresario/Inversor", "Empresa local"],
    ["Daniel Torres", "Hesvor SAS", "Empresario/Inversor", "Empresa local"],
    ["Juli\u00e1n Mauricio L\u00f3pez Mej\u00eda", "N/A", "Inversionista", "Sin fondo identificado"],
    ["Jorge Eduardo Zapata Mej\u00eda", "Golden Energy", "Empresario/Inversor", "Energ\u00eda"],
    ["Lilianne Brunstein", "Sodimac Colombia (Homecenter)", "Corporate", "Retail/Corporativo"],
    ["Juan Pablo Calder\u00f3n Bocanegra", "Due Legal", "Legal/Inversor", "Firma legal con inter\u00e9s en startups"],
]

for i, row in enumerate(independientes, 2):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 45
ws2.auto_filter.ref = f"A1:D{len(independientes)+1}"
ws2.freeze_panes = "A2"

# ========== SHEET 3: RESUMEN EJECUTIVO ==========
ws3 = wb.create_sheet("Resumen Ejecutivo")
ws3.sheet_properties.tabColor = "c8ff00"

summary_data = [
    ["RESUMEN EJECUTIVO - GoSafe AI en Startco 2026"],
    [""],
    ["M\u00e9trica", "Valor"],
    ["Total Inversionistas en plataforma", 155],
    ["VCs con fit ALTO (Healthtech/IA)", 15],
    ["VCs con fit MEDIO (Agn\u00f3stico/SaaS)", 10],
    ["Corporate VCs estrat\u00e9gicos", 7],
    ["Oportunidades adicionales", 29],
    ["Inversionistas independientes/empresarios", 56],
    [""],
    ["RONDA GoSafe", "$200,000 USD por 10% equity"],
    ["Stand en Startco", "B13 - Healthcare & Biotech"],
    ["Evento", "16-17 Abril 2026, Plaza Mayor Medell\u00edn"],
]

for i, row in enumerate(summary_data, 1):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j, value=val)
        if i == 1:
            cell.font = Font(name='Arial', size=16, bold=True, color='1a1a2e')
        elif i == 3:
            cell.font = Font(name='Arial', size=11, bold=True, color='ffffff')
            cell.fill = PatternFill('solid', fgColor='1a1a2e')
        else:
            cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(vertical='center')

ws3.column_dimensions['A'].width = 45
ws3.column_dimensions['B'].width = 40

wb.save('/sessions/nice-trusting-turing/mnt/Startco26/GoSafe_Inversionistas_Startco2026.xlsx')
print("Excel saved successfully")
